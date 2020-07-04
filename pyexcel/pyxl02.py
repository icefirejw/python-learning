import openpyxl,pprint

wb = openpyxl.load_workbook('system.xlsx')

ws = wb.active

res = {}

for row in range(2, ws.max_row+1):
    dept = ws['B'+ str(row)].value
    op_level = ws['C' + str(row)].value

    # add the default value to dictionary
    res.setdefault(dept,{'A+类':0, 'A类': 0, 'B类':0, 'C类':0, 'D类':0})

    res[dept][op_level] += 1

print(res)