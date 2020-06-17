import openpyxl

wb = openpyxl.load_workbook('01read.xlsx')

# getting sheets from workbook
#print(wb.sheetnames)
#for sheet in wb:
#    print(sheet.title)

# create a new sheet in the workbook
mysheet = wb.create_sheet('newSheet')
# print(wb.sheetnames)

# get sheet from workbook
#sheet3 = wb.get_sheet_by_name('Sheet3')
sheet4 = wb['newSheet']

# get cell from sheet
ws = wb.active
c = ws['A2']
# print(cell.value)
# get the coordinate of cell
#print("row {}, column {} is {}".format(c.row, c.column, c.value))
#print("Cell {} is {}\n".format(c.coordinate, c.value))
#print(ws.cell(row=1, column=2).value)

# row and column number is starting from "1"
#for i in range(1,8,1):
#    print(ws.cell(row=i, column=3).value)

# getting the rows and columns from sheet
colC = ws['C']
row6 = ws[6]
col_range = ws['B:C']
row_range = ws[2:6]
# print(colC)
# for col in col_range:
#     for c in col:
#         print(c.value)
# for row in ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=4):
#     for cell in row:
#         print(cell.value)

# getting the max range of sheet
# print("{} * {}".format(ws.max_row, ws.max_column))

# convert the letter to index
from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(1000))
print(column_index_from_string('AAA'))